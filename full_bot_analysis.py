#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
봇1 + 봇2 종합 거래 분석
"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl
from collections import defaultdict
import json

def load_sheet(wb, sheet_name):
    ws = wb[sheet_name]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        d = dict(zip(headers, row))
        rows.append(d)
    return rows

def safe_float(v, default=0.0):
    try:
        if v is None: return default
        return float(v)
    except:
        return default

def analyze(records, label):
    print(f"\n{'='*80}")
    print(f"  {label} — 총 {len(records)}건 분석")
    print(f"{'='*80}")

    # 분류: ENTER vs SKIP
    enters = [r for r in records if r.get('action') == 'ENTER']
    skips = [r for r in records if r.get('action') == 'SKIP']
    print(f"\n📊 기본 통계:")
    print(f"  총 기록: {len(records)} (ENTER: {len(enters)}, SKIP: {len(skips)})")

    # ENTER 중 실제 거래된 것 (tradeResult 있는 것)
    trades = [r for r in enters if r.get('tradeResult') and r['tradeResult'] in ('WIN', 'LOSS')]
    pending = [r for r in enters if not r.get('tradeResult') or r['tradeResult'] not in ('WIN', 'LOSS')]

    wins = [r for r in trades if r['tradeResult'] == 'WIN']
    losses = [r for r in trades if r['tradeResult'] == 'LOSS']

    print(f"  완료된 거래: {len(trades)} (WIN: {len(wins)}, LOSS: {len(losses)})")
    if pending:
        print(f"  미완료/진행중: {len(pending)}")

    if len(trades) == 0:
        print("  ⚠️ 완료된 거래 없음 — 상세 분석 불가")
        return trades, skips

    wr = len(wins) / len(trades) * 100
    print(f"  ✅ 승률: {wr:.1f}%")

    # PnL 분석
    total_pnl = sum(safe_float(r.get('tradePnl')) for r in trades)
    avg_pnl = total_pnl / len(trades)
    win_pnl = sum(safe_float(r.get('tradePnl')) for r in wins)
    loss_pnl = sum(safe_float(r.get('tradePnl')) for r in losses)
    avg_win = win_pnl / len(wins) if wins else 0
    avg_loss = loss_pnl / len(losses) if losses else 0

    print(f"\n💰 PnL 분석:")
    print(f"  총 PnL: ${total_pnl:.2f}")
    print(f"  평균 PnL/거래: ${avg_pnl:.4f}")
    print(f"  승리 평균: ${avg_win:.4f} | 패배 평균: ${avg_loss:.4f}")
    if avg_loss != 0:
        print(f"  Payoff Ratio: {abs(avg_win/avg_loss):.2f}:1")

    # PnL % 분석
    pnl_pcts = [safe_float(r.get('tradePnlPercent')) for r in trades if r.get('tradePnlPercent') is not None]
    if pnl_pcts:
        avg_pnl_pct = sum(pnl_pcts) / len(pnl_pcts)
        win_pnl_pcts = [safe_float(r.get('tradePnlPercent')) for r in wins if r.get('tradePnlPercent') is not None]
        loss_pnl_pcts = [safe_float(r.get('tradePnlPercent')) for r in losses if r.get('tradePnlPercent') is not None]
        avg_win_pct = sum(win_pnl_pcts) / len(win_pnl_pcts) if win_pnl_pcts else 0
        avg_loss_pct = sum(loss_pnl_pcts) / len(loss_pnl_pcts) if loss_pnl_pcts else 0
        print(f"  평균 PnL%: {avg_pnl_pct:.2f}%")
        print(f"  승리 평균%: {avg_win_pct:.2f}% | 패배 평균%: {avg_loss_pct:.2f}%")

    # 홀딩 시간
    hold_mins = [safe_float(r.get('holdingMinutes')) for r in trades if r.get('holdingMinutes')]
    if hold_mins:
        avg_hold = sum(hold_mins) / len(hold_mins)
        max_hold = max(hold_mins)
        under_60 = sum(1 for h in hold_mins if h <= 60)
        under_30 = sum(1 for h in hold_mins if h <= 30)
        print(f"\n⏱️ 홀딩 시간:")
        print(f"  평균: {avg_hold:.0f}분 | 최대: {max_hold:.0f}분")
        print(f"  30분 이내: {under_30}/{len(hold_mins)} ({under_30/len(hold_mins)*100:.0f}%)")
        print(f"  60분 이내: {under_60}/{len(hold_mins)} ({under_60/len(hold_mins)*100:.0f}%)")

    # MFE/MAE 분석
    mfes = [safe_float(r.get('maxFavorableExcursion')) for r in trades if r.get('maxFavorableExcursion') is not None]
    maes = [safe_float(r.get('maxAdverseExcursion')) for r in trades if r.get('maxAdverseExcursion') is not None]
    if mfes:
        avg_mfe = sum(mfes) / len(mfes)
        avg_mae = sum(maes) / len(maes) if maes else 0
        print(f"\n📈 MFE/MAE:")
        print(f"  평균 MFE(최대 유리): {avg_mfe:.2f}%")
        print(f"  평균 MAE(최대 불리): {avg_mae:.2f}%")

        # TP/SL 대비 MFE
        tps = [safe_float(r.get('tpPercent')) for r in trades if r.get('tpPercent')]
        sls = [safe_float(r.get('slPercent')) for r in trades if r.get('slPercent')]
        if tps:
            avg_tp = sum(tps) / len(tps)
            avg_sl = sum(sls) / len(sls) if sls else 0
            print(f"  평균 TP: {avg_tp:.2f}% | 평균 SL: {avg_sl:.2f}%")
            print(f"  MFE/TP: {avg_mfe/avg_tp:.2f}x (1.0 이상이면 TP 도달)")

    # === 방향별 분석 ===
    print(f"\n📐 방향별 분석:")
    for direction in ['Long', 'Short']:
        dt = [r for r in trades if r.get('direction') == direction]
        if not dt: continue
        dw = [r for r in dt if r['tradeResult'] == 'WIN']
        dl = [r for r in dt if r['tradeResult'] == 'LOSS']
        d_wr = len(dw) / len(dt) * 100 if dt else 0
        d_pnl = sum(safe_float(r.get('tradePnl')) for r in dt)
        print(f"  {direction}: {len(dt)}건 | WR {d_wr:.1f}% | PnL ${d_pnl:.2f}")

    # === wasReversed 분석 ===
    reversed_trades = [r for r in trades if r.get('wasReversed') == True]
    non_reversed = [r for r in trades if r.get('wasReversed') != True]
    if reversed_trades:
        rev_wins = sum(1 for r in reversed_trades if r['tradeResult'] == 'WIN')
        rev_pnl = sum(safe_float(r.get('tradePnl')) for r in reversed_trades)
        nrev_wins = sum(1 for r in non_reversed if r['tradeResult'] == 'WIN')
        nrev_pnl = sum(safe_float(r.get('tradePnl')) for r in non_reversed)
        print(f"\n🔄 Smart Reverse 효과:")
        print(f"  Reversed: {len(reversed_trades)}건, WR {rev_wins/len(reversed_trades)*100:.1f}%, PnL ${rev_pnl:.2f}")
        if non_reversed:
            print(f"  Non-Rev:  {len(non_reversed)}건, WR {nrev_wins/len(non_reversed)*100:.1f}%, PnL ${nrev_pnl:.2f}")

    # === 세션별 분석 ===
    print(f"\n🕐 세션별 분석:")
    sessions = defaultdict(list)
    for r in trades:
        s = r.get('session') or 'UNKNOWN'
        sessions[s].append(r)
    for s, st in sorted(sessions.items(), key=lambda x: -len(x[1])):
        sw = sum(1 for r in st if r['tradeResult'] == 'WIN')
        s_pnl = sum(safe_float(r.get('tradePnl')) for r in st)
        s_wr = sw / len(st) * 100
        print(f"  {s:20s}: {len(st):3d}건 | WR {s_wr:5.1f}% | PnL ${s_pnl:8.2f}")

    # === 레짐별 분석 ===
    print(f"\n🧬 레짐별 분석:")
    regimes = defaultdict(list)
    for r in trades:
        reg = r.get('regime') or 'UNKNOWN'
        regimes[reg].append(r)
    for reg, rt in sorted(regimes.items(), key=lambda x: -len(x[1])):
        rw = sum(1 for r in rt if r['tradeResult'] == 'WIN')
        r_pnl = sum(safe_float(r.get('tradePnl')) for r in rt)
        r_wr = rw / len(rt) * 100
        print(f"  {reg:30s}: {len(rt):3d}건 | WR {r_wr:5.1f}% | PnL ${r_pnl:8.2f}")

    # === Zone Type별 분석 ===
    print(f"\n🎯 Zone Type별 분석:")
    zones = defaultdict(list)
    for r in trades:
        z = r.get('zoneType') or r.get('strategy') or 'UNKNOWN'
        zones[z].append(r)
    for z, zt in sorted(zones.items(), key=lambda x: -len(x[1])):
        zw = sum(1 for r in zt if r['tradeResult'] == 'WIN')
        z_pnl = sum(safe_float(r.get('tradePnl')) for r in zt)
        z_wr = zw / len(zt) * 100
        print(f"  {z:25s}: {len(zt):3d}건 | WR {z_wr:5.1f}% | PnL ${z_pnl:8.2f}")

    # === 레버리지별 분석 ===
    print(f"\n⚡ 레버리지별 분석:")
    levs = defaultdict(list)
    for r in trades:
        lev = r.get('leverage')
        if lev is not None:
            lev = int(safe_float(lev))
            bucket = f"{lev}x"
        else:
            bucket = 'N/A'
        levs[bucket].append(r)
    for lev, lt in sorted(levs.items(), key=lambda x: -len(x[1])):
        lw = sum(1 for r in lt if r['tradeResult'] == 'WIN')
        l_pnl = sum(safe_float(r.get('tradePnl')) for r in lt)
        l_wr = lw / len(lt) * 100
        print(f"  {lev:8s}: {len(lt):3d}건 | WR {l_wr:5.1f}% | PnL ${l_pnl:8.2f}")

    # === exitReason 분석 ===
    print(f"\n🚪 Exit Reason 분석:")
    exits = defaultdict(list)
    for r in trades:
        ex = r.get('exitReason') or 'UNKNOWN'
        exits[ex].append(r)
    for ex, et in sorted(exits.items(), key=lambda x: -len(x[1])):
        ew = sum(1 for r in et if r['tradeResult'] == 'WIN')
        e_pnl = sum(safe_float(r.get('tradePnl')) for r in et)
        e_wr = ew / len(et) * 100
        print(f"  {ex:30s}: {len(et):3d}건 | WR {e_wr:5.1f}% | PnL ${e_pnl:8.2f}")

    # === SKIP 분석 ===
    if skips:
        print(f"\n⏭️ SKIP 분석 (총 {len(skips)}건):")
        skip_reasons = defaultdict(int)
        for r in skips:
            sr = r.get('skipReason') or 'UNKNOWN'
            skip_reasons[sr] += 1
        for sr, cnt in sorted(skip_reasons.items(), key=lambda x: -x[1]):
            print(f"  {sr:35s}: {cnt:4d}건 ({cnt/len(skips)*100:.1f}%)")

        # SKIP 이후 시장 움직임 (moveAfter5min)
        skip_moves = [(r.get('skipReason'), safe_float(r.get('moveAfter5min')), safe_float(r.get('moveAfter15min')))
                      for r in skips if r.get('moveAfter5min') is not None]
        if skip_moves:
            print(f"\n  📉 SKIP 후 시장 움직임 (놓친 기회 분석):")
            skip_by_reason = defaultdict(list)
            for sr, m5, m15 in skip_moves:
                skip_by_reason[sr or 'UNKNOWN'].append((m5, m15))
            for sr, moves in sorted(skip_by_reason.items(), key=lambda x: -len(x[1])):
                avg5 = sum(m[0] for m in moves) / len(moves)
                avg15 = sum(m[1] for m in moves) / len(moves)
                print(f"    {sr:33s}: avg5m={avg5:+.2f}%, avg15m={avg15:+.2f}% ({len(moves)}건)")

    # === 티커별 분석 (상위 15) ===
    print(f"\n🏆 티커별 분석 (거래 수 상위 15):")
    tickers = defaultdict(list)
    for r in trades:
        t = r.get('ticker') or 'UNKNOWN'
        tickers[t].append(r)
    sorted_tickers = sorted(tickers.items(), key=lambda x: -len(x[1]))[:15]
    for t, tt in sorted_tickers:
        tw = sum(1 for r in tt if r['tradeResult'] == 'WIN')
        t_pnl = sum(safe_float(r.get('tradePnl')) for r in tt)
        t_wr = tw / len(tt) * 100
        print(f"  {t:18s}: {len(tt):3d}건 | WR {t_wr:5.1f}% | PnL ${t_pnl:8.4f}")

    # === 연속 패/승 분석 ===
    streaks_loss = []
    streaks_win = []
    cur_streak = 0
    cur_type = None
    for r in trades:
        res = r['tradeResult']
        if res == cur_type:
            cur_streak += 1
        else:
            if cur_type == 'LOSS' and cur_streak > 0:
                streaks_loss.append(cur_streak)
            elif cur_type == 'WIN' and cur_streak > 0:
                streaks_win.append(cur_streak)
            cur_streak = 1
            cur_type = res
    if cur_type == 'LOSS':
        streaks_loss.append(cur_streak)
    elif cur_type == 'WIN':
        streaks_win.append(cur_streak)

    print(f"\n📊 연속 스트릭:")
    if streaks_loss:
        print(f"  최대 연패: {max(streaks_loss)} | 평균 연패: {sum(streaks_loss)/len(streaks_loss):.1f}")
    if streaks_win:
        print(f"  최대 연승: {max(streaks_win)} | 평균 연승: {sum(streaks_win)/len(streaks_win):.1f}")

    # === TP/SL RR ratio vs 결과 ===
    print(f"\n📏 Risk/Reward 분석:")
    rr_buckets = defaultdict(list)
    for r in trades:
        tp = safe_float(r.get('tpPercent'))
        sl = safe_float(r.get('slPercent'))
        if sl > 0 and tp > 0:
            rr = tp / sl
            if rr < 1.0:
                bucket = '<1:1'
            elif rr < 1.5:
                bucket = '1-1.5:1'
            elif rr < 2.0:
                bucket = '1.5-2:1'
            elif rr < 3.0:
                bucket = '2-3:1'
            else:
                bucket = '3+:1'
            rr_buckets[bucket].append(r)
    for bucket in ['<1:1', '1-1.5:1', '1.5-2:1', '2-3:1', '3+:1']:
        bt = rr_buckets.get(bucket, [])
        if not bt: continue
        bw = sum(1 for r in bt if r['tradeResult'] == 'WIN')
        b_pnl = sum(safe_float(r.get('tradePnl')) for r in bt)
        b_wr = bw / len(bt) * 100
        print(f"  RR {bucket:10s}: {len(bt):3d}건 | WR {b_wr:5.1f}% | PnL ${b_pnl:8.2f}")

    # === Equity 추이 ===
    equities = [safe_float(r.get('totalEquityAtEntry')) for r in trades if r.get('totalEquityAtEntry')]
    if equities:
        print(f"\n💼 Equity 추이:")
        print(f"  시작: ${equities[0]:.2f} → 최근: ${equities[-1]:.2f}")
        print(f"  최고: ${max(equities):.2f} | 최저: ${min(equities):.2f}")
        if equities[0] > 0:
            print(f"  변동: {((equities[-1] - equities[0]) / equities[0]) * 100:.1f}%")

    # === BTC 추세별 분석 ===
    btc_trades = [r for r in trades if r.get('btcTrend')]
    if btc_trades:
        print(f"\n₿ BTC 추세별 분석:")
        btc_groups = defaultdict(list)
        for r in btc_trades:
            btc_groups[r['btcTrend']].append(r)
        for bt, bg in sorted(btc_groups.items(), key=lambda x: -len(x[1])):
            bw = sum(1 for r in bg if r['tradeResult'] == 'WIN')
            b_pnl = sum(safe_float(r.get('tradePnl')) for r in bg)
            b_wr = bw / len(bg) * 100
            print(f"  BTC {bt:6s}: {len(bg):3d}건 | WR {b_wr:5.1f}% | PnL ${b_pnl:8.2f}")

    return trades, skips


# ===== MAIN =====
wb = openpyxl.load_workbook(r'C:\Users\michj\Downloads\거래내역 통계.xlsx', data_only=True)

bot2_records = load_sheet(wb, 'Bot2')
bot1_records = load_sheet(wb, 'Bot1')

print("🤖 봇 2개 거래 데이터 종합 분석")
print(f"Bot2: {len(bot2_records)}건 | Bot1: {len(bot1_records)}건")

trades2, skips2 = analyze(bot2_records, "🖥️ Bot2 (맥북/크롬)")
trades1, skips1 = analyze(bot1_records, "💻 Bot1 (HP/엣지)")

# === 봇 간 비교 ===
print(f"\n{'='*80}")
print(f"  🔀 Bot1 vs Bot2 비교")
print(f"{'='*80}")

all_trades = trades1 + trades2
if all_trades:
    all_wins = sum(1 for r in all_trades if r['tradeResult'] == 'WIN')
    all_pnl = sum(safe_float(r.get('tradePnl')) for r in all_trades)
    print(f"\n  합산: {len(all_trades)}거래 | WR {all_wins/len(all_trades)*100:.1f}% | PnL ${all_pnl:.2f}")

# 공통 종목 분석
tickers1 = set(r.get('ticker') for r in trades1)
tickers2 = set(r.get('ticker') for r in trades2)
common = tickers1 & tickers2
only1 = tickers1 - tickers2
only2 = tickers2 - tickers1
print(f"\n  공통 종목: {len(common)}개")
print(f"  Bot1만: {len(only1)}개 | Bot2만: {len(only2)}개")

# === 전체 합산 핵심 문제 식별 ===
print(f"\n{'='*80}")
print(f"  🔍 핵심 문제 식별 (전체 합산)")
print(f"{'='*80}")

if all_trades:
    # 1. 가장 많이 잃은 레짐
    regime_pnl = defaultdict(float)
    regime_cnt = defaultdict(int)
    for r in all_trades:
        reg = r.get('regime') or 'UNKNOWN'
        regime_pnl[reg] += safe_float(r.get('tradePnl'))
        regime_cnt[reg] += 1
    worst_regimes = sorted(regime_pnl.items(), key=lambda x: x[1])[:3]
    print(f"\n  🚨 최악 레짐 (PnL 하위 3):")
    for reg, pnl in worst_regimes:
        cnt = regime_cnt[reg]
        wr = sum(1 for r in all_trades if r.get('regime') == reg and r['tradeResult'] == 'WIN') / cnt * 100
        print(f"    {reg}: ${pnl:.2f} ({cnt}건, WR {wr:.1f}%)")

    # 2. 가장 많이 잃은 세션
    sess_pnl = defaultdict(float)
    sess_cnt = defaultdict(int)
    for r in all_trades:
        s = r.get('session') or 'UNKNOWN'
        sess_pnl[s] += safe_float(r.get('tradePnl'))
        sess_cnt[s] += 1
    worst_sessions = sorted(sess_pnl.items(), key=lambda x: x[1])[:3]
    print(f"\n  🚨 최악 세션 (PnL 하위 3):")
    for s, pnl in worst_sessions:
        cnt = sess_cnt[s]
        wr = sum(1 for r in all_trades if r.get('session') == s and r['tradeResult'] == 'WIN') / cnt * 100
        print(f"    {s}: ${pnl:.2f} ({cnt}건, WR {wr:.1f}%)")

    # 3. 가장 많이 잃은 Zone
    zone_pnl = defaultdict(float)
    zone_cnt = defaultdict(int)
    for r in all_trades:
        z = r.get('zoneType') or r.get('strategy') or 'UNKNOWN'
        zone_pnl[z] += safe_float(r.get('tradePnl'))
        zone_cnt[z] += 1
    worst_zones = sorted(zone_pnl.items(), key=lambda x: x[1])[:3]
    print(f"\n  🚨 최악 Zone (PnL 하위 3):")
    for z, pnl in worst_zones:
        cnt = zone_cnt[z]
        wr = sum(1 for r in all_trades if (r.get('zoneType') or r.get('strategy')) == z and r['tradeResult'] == 'WIN') / cnt * 100
        print(f"    {z}: ${pnl:.2f} ({cnt}건, WR {wr:.1f}%)")

    # 4. TP 도달률 vs MFE
    tp_miss = 0
    tp_hit = 0
    mfe_above_tp = []
    for r in all_trades:
        mfe = safe_float(r.get('maxFavorableExcursion'))
        tp = safe_float(r.get('tpPercent'))
        if tp > 0 and mfe > 0:
            if mfe >= tp:
                tp_hit += 1
            else:
                tp_miss += 1
                mfe_above_tp.append(mfe / tp)
    if tp_hit + tp_miss > 0:
        print(f"\n  🎯 TP 도달 분석:")
        print(f"    TP 도달: {tp_hit}/{tp_hit+tp_miss} ({tp_hit/(tp_hit+tp_miss)*100:.1f}%)")
        print(f"    TP 미도달: {tp_miss}건")
        if mfe_above_tp:
            avg_reach = sum(mfe_above_tp) / len(mfe_above_tp)
            print(f"    미도달 시 평균 MFE/TP: {avg_reach:.1f}x (TP의 {avg_reach*100:.0f}%까지 갔다가 반전)")

    # 5. 연패 후 큰 손실
    big_losses = [r for r in all_trades if safe_float(r.get('tradePnl')) < -0.5]
    if big_losses:
        print(f"\n  💀 큰 손실 거래 (>${0.5} 손실): {len(big_losses)}건")
        for r in sorted(big_losses, key=lambda x: safe_float(x.get('tradePnl')))[:5]:
            print(f"    {r.get('ticker')}: ${safe_float(r.get('tradePnl')):.4f} | {r.get('regime')} | {r.get('session')} | Lev {r.get('leverage')}x | {r.get('exitReason')}")

print("\n\n✅ 분석 완료!")
